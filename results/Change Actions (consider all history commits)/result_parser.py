# -*- coding: utf-8 -*-
"""
@author: Administrator
"""


# borrow from parseTxt.py
all_projs = ["Chart", "Math", "Lang", "Mockito", "Time", "Closure"]
all_folders = ["Chart", "Math", "Lang", "Mockito", "Time", "Closure"]
all_numbers = [26, 106, 65, 38, 27, 133]
d4j_dict = {}
for i in range(0, len(all_projs)):
    d4j_dict[all_projs[i]] = all_numbers[i]
    
import os
from Tree import CCITree
    
class Bug:
    """
    describe the bug project, id, and repair time, etc.
    """
    def __init__(self, proj, id, match_list = [], cci_list = [], 
                 all_commits_len = -1, repaired = "unknown", correct="", tree_list = [], depth = 0):
        self.proj = proj
        self.id = id
        self.match_list = match_list #[freq, cci_no, rank]
        self.cci_list = cci_list #
        self.correct = correct
        self.repaired = repaired
        self.all_commits_len = all_commits_len
        self.tree_list = tree_list
        self.depth = depth
    
    def toString(self):
        return "{}_{}_{}".format(self.proj, self.id, self.correct)
    
    def toString2(self):
        return "{}_{}".format(self.proj, self.id)
    
    def __eq__(self, other):
        return self.proj.capitalize() == other.proj.capitalize() and str(self.id) == str(other.id)


def check_commitNo(cci_path):
    import os, re
    commitNoCnt = 0
    for file in os.listdir(cci_path):
        if re.match("CommitId-", file):
            commitNoCnt += 1
    if commitNoCnt > 1:
        print("{} have more than one commitId".format(cci_path))

def get_CCI_info(path):
    bugs = []
    for i in range(0, len(all_folders)):
        folder = all_folders[i]
        proj = all_projs[i]
        for id_name in range(1, d4j_dict[proj]+1):
            cci_path = path + "\\{}-exp1\\{}\\".format(folder, id_name) # str
            if os.path.exists(cci_path):
                parse_CCI(cci_path, bugs, proj, id_name, 1)
            
            cci_path = path + "\\{}-exp2\\{}\\".format(folder, id_name) # str
            if os.path.exists(cci_path):
                parse_CCI(cci_path, bugs, proj, id_name, 2)

            cci_path = path + "\\{}\\{}\\".format(folder, id_name) # str
            parse_CCI(cci_path, bugs, proj, id_name, 0)
            
    return bugs

def parse_CCI(cci_path, bugs, proj, id_name, exp_index):
    check_commitNo(cci_path)

    if not hasCommitID(cci_path):
        print("{}_{}_{} has no commitID".format(proj, id_name, exp_index))
        return 
    
    freq_file = cci_path + "FixPatternFreqs.txt"
    if not os.path.exists(freq_file):
            print("{}_{}_{} has no FixPatternFreqs.txt".format(proj, id_name, exp_index))
            return
    freq_list, all_commits_len = parse_freq(freq_file)
    
    match_result_file = cci_path + "matchResult.txt"
    if not os.path.exists(match_result_file):
            print("{}_{}_{} has no match_result_file".format(proj, id_name, exp_index))
            return
    match_list = get_match_result(match_result_file, freq_list)
    
    cci_file = cci_path + "CCI"
    if not os.path.exists(cci_file):
        cci_file = cci_path + "changeAction.txt"   # for Chart
        if not os.path.exists(cci_file):
            print("{}_{}_{} has no CCI".format(proj, id_name, exp_index))
            return
    cci_list = parse_CCI_file(cci_file)
    
    tree_list = []
    depth = -1
    depth_list = []
    once_del = False
    for cci in cci_list:
        del_op = False
        cci_tree = CCITree()
        nodes = []
        op_cnt = 0
        for op in cci:
            space_cnt = op.split(":")[1].split("(")[0].count(" ")
            level = int(space_cnt/3.0)
            
            if level > depth:
                depth = level
            
            op_index = int(op.split(":")[0][2:]) - 1
            sub_op1 = op.split("(")[1].split(",")[0].strip()
            
            if sub_op1 == "DEL" and op_cnt == 0:
                del_op = True
                once_del = True
            op_cnt += 1
            
            sub_op2 = op.split("(")[1].split(",")[1].strip()
            element = sub_op1 + " " + sub_op2
            
            parent_str = op.split("(")[1].split(",")[2].strip()
            if parent_str != "null":
                parent_index = int(parent_str[2:]) - 1 #bug fix 
            else:
                parent_index = -1
            
            children_str = op.split("(")[1].split(",")[3].strip().split(")")[0]
            children_index_list = []
            for child in children_str.split(" "):
                if child != "null":
                    child_index = int(child[2:]) - 1
                    children_index_list.append(child_index)
            
            node = CCITree.Node(element, space_cnt = space_cnt, op_index = op_index, 
                                parent_index = parent_index, children_index_list = children_index_list, children = [] )
            
            nodes.append(node)
            
        for node in nodes:
            if node.space_cnt == 0:
                cci_tree._root = node
                for index in node.children_index_list:
                    cci_tree.add_child_node(nodes[index], node)
            else:
                for index in node.children_index_list:
                    cci_tree.add_child_node(nodes[index], node)
                node.parent = nodes[node.parent_index]  # so hard bug 
            
        tree_list.append(cci_tree)
        
        if del_op == True:
            depth = 0
        
        depth_list.append(depth)
        
    # check depth from 2 ways
    depth2 = get_depth(tree_list)
    if once_del == True and max(depth_list) == 0:
        print(" proj_id: {}_{}, depth:{}".format(proj, id_name, max(depth_list)))

#self, proj, id, match_list = [], cci_list = [], all_commits_len = -1, repaired = "unknown", correct="", tree_list = [], depth = 0):
    bug = Bug(proj = proj, id = id_name, match_list = match_list, cci_list = cci_list, 
              all_commits_len = all_commits_len, tree_list = tree_list, depth = max(depth_list))
    if bug not in bugs:
        bugs.append(bug)
    else:
        print("repeated in {}_{}_{}".format(proj, id_name, exp_index))

def get_depth(tree_list):
    depth = 0
    for tree in tree_list:
        cur_depth = tree.height(tree.make_position(tree._root))
        if cur_depth > depth:
            depth = cur_depth
    return depth

def hasCommitID(cci_path):
    import os
    import re
    
    for file in os.listdir(cci_path):
        name = os.path.basename(file)
        if re.match("CommitId-", name):
            return True
    return False

def make_sheet(path, bugs, sheet_name = 'newSheet'):
        import xlsxwriter
        # Create a workbook and add a worksheet.
        workbook = xlsxwriter.Workbook(path)
        worksheet = workbook.add_worksheet(sheet_name)
        
        # Add a bold format to use to highlight cells.
        title_f = workbook.add_format()
        title_f.set_font_size(15)
        title_f.set_align('center')
        title_f.set_align('vcenter')
        title_f.set_bold(True)
        title_f.set_bg_color('orange')
        title_f.set_font('Times New Roman')
        
        cell_f = workbook.add_format()
        cell_f.set_font_size(11)
        cell_f.set_align('center')
        cell_f.set_align('vcenter')
        cell_f.set_bold(False)
        cell_f.set_font('Times New Roman')
        cell_f.set_text_wrap(True)

        worksheet.write(0, 0,'bug_id',title_f)
        worksheet.write(0, 1,'number of CCIs',title_f)
        worksheet.write(0, 2,'matched number of CCIs',title_f)
        worksheet.write(0, 3,'matched CCIs Index',title_f)
        worksheet.write(0, 4,'frequencies',title_f)
        worksheet.write(0, 5,'rank',title_f)
        worksheet.write(0, 6,'number of all commits',title_f)
        worksheet.write(0, 7,'if correctly fixed',title_f)
        
        worksheet.set_column(0,0,20)
        worksheet.set_column(1,7,33)
        
        row = 1
        for proj in ["Chart", "Closure", "Lang", "Math", "Mockito", "Time"]:
            for id_name in range(1, d4j_dict[proj] +1): #bug fix: forget to add range
                bug = Bug(proj, id_name)
                worksheet.write(row, 0, bug.toString2(), cell_f)
                if bug in bugs:
                    index = bugs.index(bug)
                    
                    target_bug = bugs[index]
                    match_list = target_bug.match_list 
                    
                    freq = ""
                    cci_no = ""
                    rank = ""
                    for match in match_list:
                        freq += " " + str(match[0])
                        cci_no += " " + str(match[1])
                        rank += " " + str(match[2])
                    
                    # for recording depth
                    worksheet.write(row, 1, target_bug.depth, cell_f)
                    
                    worksheet.write(row, 2, len(match_list), cell_f)
                    worksheet.write(row, 3, cci_no, cell_f)
                    worksheet.write(row, 4, freq, cell_f)
                    worksheet.write(row, 5, rank, cell_f)
                    worksheet.write(row, 6, target_bug.all_commits_len, cell_f)
                    worksheet.write(row, 7, target_bug.repaired, cell_f)
                row += 1 #bug fix. forget to add 1
        workbook.close()
            
def parse_freq(freq_file):
    import re
    with open(freq_file, encoding = 'utf-8', mode = 'r') as f:
        lines = f.readlines()
        freq_list = []
        all_commits_list = []
        for line in lines:
            line = line.strip()
            if re.match("Freq: ", line):
                freq_list.append(int(line.split(" ")[1])) # get freq
            if re.match("CommitIds: ", line):
                commits = line.split(" ")
                for i in range(1,len(commits)): #discard first element. "i.e., CommitIds-"
                    if commits[i] not in all_commits_list:
                        all_commits_list.append(commits[i])
        return freq_list, len(all_commits_list)
                
def get_match_result(match_result_file, freq_list):
    import re
    with open(match_result_file, encoding = 'utf-8', mode = 'r') as f:
        lines = f.readlines()
        freq = -1
        cci_no = -1
        match_list = []
        for line in lines:
            line = line.strip()
            if re.match("No.", line):
                cci_no = int(line.split(" ")[0].split(".")[1]) + 1
            if re.match("freq: ", line):
                freq = int(line.split(" ")[1])
            if freq != -1 and cci_no != -1:
                rank  = get_rank(freq, freq_list)
                match_list.append([freq, cci_no, rank])
                freq = -1
                cci_no = -1 # re-init
        return match_list
            
def get_rank(freq, freq_list):
    for i in range(0, len(freq_list)):
        if freq >= freq_list[i]:
            return i+1
        
def parse_CCI_file(cci_file):
    with open(cci_file, encoding = 'utf-8', mode = 'r') as f:
        lines = f.readlines()
        cci_list = []
        cci = []
        for line in lines:
            line = line.strip() # wipe out \n
            if len(line) != 0:
                cci.append(line) #add an op
            elif len(cci) != 0: 
                cci_list.append(cci)
                cci = [] #bug fix: re-init
                
        return cci_list
                
    
def get_bugs_info(bugs):
    freq = 0
    rank = 0
    top10_rank_cnt = 0
    not_empty_cnt = 0
    full_match_cnt = 0
    within_10_full_cnt = 0
    within_10_partial_cnt = 0
    
    CCI_no_dict = {}
    
    for bug in bugs:
        match_list = bug.match_list
        
        cci_no = len(bug.cci_list)
        if cci_no != 0:
            if cci_no in CCI_no_dict.keys():
                value = CCI_no_dict[cci_no]
                CCI_no_dict[cci_no] = value + 1
            else:
                CCI_no_dict[cci_no] = 1
        
        if len(bug.cci_list) == len(match_list):
            full_match_cnt += 1
            
        if len(match_list) != 0:
            not_empty_cnt += 1
        
            within_10_full = True
            within_10_partial = False
            for match in match_list:
                freq += int(match[0])
                rank += int(match[2])
                if int(match[2]) <= 10:
                    top10_rank_cnt += 1
                    within_10_partial = True
                else:
                    within_10_full = False
            if within_10_partial == True:
                within_10_partial_cnt += 1
            if within_10_full == True:
                within_10_full_cnt += 1
            
    print("bugs size: {}, not empty: {}".format(len(bugs), not_empty_cnt))
    print("freq average: {}".format(freq/not_empty_cnt))
    print("rank average: {}".format(rank/not_empty_cnt))
    print("number of CCIs with freq ranking within top10: {}".format(top10_rank_cnt))
    print("number of bugs with all freqs ranking within top10: {}".format(within_10_full_cnt))
    print("number of bugs with partial freqs ranking within top10: {}".format(within_10_partial_cnt))
    print("number of fully matched bugs: {}".format(full_match_cnt))
    
    CCI_no_dict_sort = [(k, CCI_no_dict[k]) for k in sorted(CCI_no_dict.keys())]
    return CCI_no_dict_sort
        
    

#read excel
def read_excel(bugs, path = r"E:\APR Tools.xlsx"):
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
     
    for row_cnt in range(2, 397):
        bug_id = ws.cell(row = row_cnt, column = 1).value
        current_bug = Bug(bug_id.split("_")[0], bug_id.split("_")[1])
        if current_bug in bugs:
            index = bugs.index(current_bug)
            target_bug = bugs[index]
            
            #judge if repaired
            yes_flag = False
            for column_cnt in range(2, 33):
                result = ws.cell(row = row_cnt, column = column_cnt).value
                if result == "✓":
                    target_bug.repaired = "yes"
                    yes_flag = True
                    break
            if yes_flag == False:
                target_bug.repaired = "no"
    
def read_excel_each_apr(path = r"E:\each-apr.xlsx"):
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active  
    
    apr_dict = {}
    for column_cnt in range(3, 27):#36
        apr_name = ws.cell(row = 1, column = column_cnt).value
        print("\n\n=================  {} ========================".format(apr_name))
        apr_name_dict = {}
        for row_cnt in range(2, 397):
            bug_id = ws.cell(row = row_cnt, column = 1).value
            cci_number = int(ws.cell(row = row_cnt, column = 2).value)
            result = ws.cell(row = row_cnt, column = column_cnt).value
            if result == "✓":
                repair = "correct"
                if cci_number not in apr_name_dict.keys():
                    apr_name_dict[cci_number] = [1, 0, 0]
                else:
                    cci_no_list = apr_name_dict[cci_number]
                    apr_name_dict[cci_number] = [cci_no_list[0] + 1, cci_no_list[1], cci_no_list[2]]
            elif result == "✗":
                repair = "incorrect"
                if cci_number not in apr_name_dict.keys():
                    apr_name_dict[cci_number] = [0, 1, 0]
                else:
                    cci_no_list = apr_name_dict[cci_number]
                    apr_name_dict[cci_number] = [cci_no_list[0], cci_no_list[1] + 1, cci_no_list[2]]
            else:
                repair = "unfixed"
                if cci_number not in apr_name_dict.keys():
                    apr_name_dict[cci_number] = [0, 0, 1]
                else:
                    cci_no_list = apr_name_dict[cci_number]
                    apr_name_dict[cci_number] = [cci_no_list[0], cci_no_list[1], cci_no_list[2] + 1]
            
            # print
            if cci_number == 1 and repair == "correct":
                print("cci_no: {}, bug id: {}, repair result: {}".format(cci_number, bug_id, repair))
                
        list1 = list(apr_name_dict.keys())
        list1.sort()
        apr_name_dict_sort = {}
        for i in list1:
            for k in apr_name_dict:
                if i == k:
                    apr_name_dict_sort.setdefault(i, apr_name_dict[k])
        
        apr_dict[apr_name] = apr_name_dict_sort
        
        print_info(apr_dict)
    return apr_dict

def print_info(apr_dict):
    # for cci no
    for cci_no in range(0, 30):
        print("----------- Depth of change actions: {} ---------------------------".format(cci_no))
        for apr_name in apr_dict:
            apr_name_dict = apr_dict[apr_name]
            if cci_no in apr_name_dict.keys():
                if apr_name_dict[cci_no][0] > 0:
                    print(apr_name)

# borrow from parseTxt.py
def make_sheet_each_apr(path, apr_dict, sheet_name = ''):
        import xlsxwriter
        
        cnt = 1
        workbook = xlsxwriter.Workbook(path)
        for apr_name in apr_dict.keys():
            apr_name_dict = apr_dict[apr_name]
            
            # Create a workbook and add a worksheet.
            worksheet = workbook.add_worksheet(sheet_name + " " + apr_name[0:6] + " " + str(cnt))
            cnt += 1
            
            # Add a bold format to use to highlight cells.
            title_f = workbook.add_format()
            title_f.set_font_size(15)
            title_f.set_align('center')
            title_f.set_align('vcenter')
            title_f.set_bold(True)
            title_f.set_bg_color('orange')
            title_f.set_font('Times New Roman')
            
            cell_f = workbook.add_format()
            cell_f.set_font_size(11)
            cell_f.set_align('center')
            cell_f.set_align('vcenter')
            cell_f.set_bold(False)
            cell_f.set_font('Times New Roman')
            cell_f.set_text_wrap(True)
            

            worksheet.write(0, 0,'cci_no',title_f)
            worksheet.write(0, 1,"correct",title_f)
            worksheet.write(0, 2,"incorrect",title_f)
            worksheet.write(0, 3,'unfixed',title_f)
            worksheet.set_column(0,3,20)
            
            row = 1
            for cci_no in apr_name_dict.keys():
                cci_no_list = apr_name_dict[cci_no]
                worksheet.write(row, 0, cci_no, cell_f)
                worksheet.write(row, 1, cci_no_list[0], cell_f)
                worksheet.write(row, 2, cci_no_list[1], cell_f)
                worksheet.write(row, 3, cci_no_list[2], cell_f)
                row += 1 #bug fix. forget to add 1
        workbook.close()
            
if __name__ == "__main__":
    path = r".\\"
    bugs = get_CCI_info(path)
    
    ws = read_excel(bugs)
    
    CCI_no_dict_sort = get_bugs_info(bugs)
    
    path2 = r".\result-consider-all-commits.xlsx"
    make_sheet(path2, bugs)
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    